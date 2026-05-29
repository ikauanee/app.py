from flask import Flask, request, jsonify, render_template_string
import pandas as pd
import unicodedata
import io
import base64
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

UNIDADES_ALVO = [
    "ALTA FLORESTA",
    "CANARANA", 
    "COLIDER",
    "COMODORO",
    "PIQUETE",
    "PONTES E LACERDA",
    "ARAGUAIA",
    "PALESTINA",
    "ITAPOA",
    "SAO GABRIEL"
]

NOMES_PLANILHA = {
    "ALTA FLORESTA": "ÁGUAS ALTA FLORESTA LTDA",
    "CANARANA": "ÁGUAS CANARANA LTDA",
    "COLIDER": "ÁGUAS COLIDER LTDA",
    "COMODORO": "ÁGUAS COMODORO LTDA",
    "PIQUETE": "ÁGUAS PIQUETE S.A.",
    "PONTES E LACERDA": "ÁGUAS PONTES E LACERDA LTDA",
    "ARAGUAIA": "ARAGUAIA SANEAMENTO S.A - RDN",
    "PALESTINA": "EMPRESA DE SANEAMENTO DE PALESTINA - ESAP S/A",
    "ITAPOA": "ITAPOÁ SANEAMENTO LTDA.",
    "SAO GABRIEL": "SÃO GABRIEL SANEAMENTO S.A"
}

VALORES_UNITARIOS = {
    "franquia": 450.00,
    "mensagens_receptivas": 0.11,
    "licencas": 90.00,
    "atendimento_presencial": 1000.00,
    "whatsapp_marketing": 0.48,
    "whatsapp_utility": 0.14,
    "numero_oficial_waba": 400.00
}

ALIASES = {
    "ALTA FLORESTA": "ALTA FLORESTA",
    "ARAGUAIA": "ARAGUAIA",
    "CANARANA": "CANARANA",
    "COLIDER": "COLIDER", "COLIDOR": "COLIDER", "COLÍDER": "COLIDER",
    "COMODORO": "COMODORO",
    "ITAPOA": "ITAPOA", "ITAPOÃ": "ITAPOA", "ITAPOA/SAO JOSE": "ITAPOA",
    "PALESTINA": "PALESTINA", "ESAP": "PALESTINA",
    "PIQUETE": "PIQUETE", "PIQUETÉ": "PIQUETE",
    "PONTES E LACERDA": "PONTES E LACERDA", "PONTES LACERDA": "PONTES E LACERDA",
    "SAO GABRIEL": "SAO GABRIEL", "SÃO GABRIEL": "SAO GABRIEL",
    "HIDROFORTE": "HIDROFORTE",
}

def normalizar(texto):
    if pd.isna(texto):
        return ""
    texto_nfd = unicodedata.normalize('NFD', str(texto))
    texto_limpo = "".join(c for c in texto_nfd if unicodedata.category(c) != 'Mn')
    return " ".join(texto_limpo.upper().strip().split())

def identificar_unidade(texto):
    t = normalizar(texto)
    if not t:
        return None
    for chave in sorted(ALIASES.keys(), key=len, reverse=True):
        if normalizar(chave) in t:
            return ALIASES[chave]
    return None

def super_diagnostico(caminho, nome_arquivo):
    """Diagnóstico COMPLETO do arquivo"""
    print(f"\n{'='*80}")
    print(f"🔬 SUPER DIAGNÓSTICO: {nome_arquivo}")
    print(f"{'='*80}")
    
    if not caminho or not os.path.exists(caminho):
        print("❌ Arquivo não encontrado")
        return None
    
    tamanho = os.path.getsize(caminho)
    print(f"📏 Tamanho: {tamanho} bytes")
    
    # 1. Mostrar conteúdo bruto
    print(f"\n📄 CONTEÚDO BRUTO (primeiras 5 linhas):")
    for encoding in ['utf-8', 'latin1', 'cp1252']:
        try:
            with open(caminho, 'r', encoding=encoding) as f:
                linhas = [f.readline().strip() for _ in range(5)]
            print(f"   Encoding: {encoding}")
            for i, linha in enumerate(linhas, 1):
                print(f"   Linha {i}: {linha[:300]}")
            break
        except:
            continue
    
    # 2. Tentar ler com pandas - TESTAR TODAS AS COMBINAÇÕES
    print(f"\n🔧 TESTANDO LEITURA COM PANDAS:")
    
    df = None
    config_sucesso = None
    
    for sep in [',', ';', '\t', '|']:
        for encoding in ['utf-8-sig', 'utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
            try:
                df_teste = pd.read_csv(caminho, encoding=encoding, sep=sep, nrows=3)
                if len(df_teste.columns) > 1:
                    print(f"   ✅ SUCESSO: sep='{sep}', encoding='{encoding}'")
                    print(f"      Shape: {df_teste.shape}")
                    print(f"      Colunas: {list(df_teste.columns)}")
                    df = df_teste
                    config_sucesso = {'sep': sep, 'encoding': encoding}
                    break
            except:
                continue
        if df is not None:
            break
    
    if df is None:
        print("   ❌ NÃO FOI POSSÍVEL LER O ARQUIVO")
        return None
    
    # 3. Ler arquivo completo
    try:
        df_completo = pd.read_csv(
            caminho, 
            encoding=config_sucesso['encoding'], 
            sep=config_sucesso['sep']
        )
        print(f"\n✅ Arquivo completo: {len(df_completo)} linhas, {len(df_completo.columns)} colunas")
    except Exception as e:
        print(f"❌ Erro ao ler completo: {e}")
        return None
    
    # 4. Análise detalhada de CADA coluna
    print(f"\n🔍 ANÁLISE DETALHADA DE CADA COLUNA:")
    print(f"   (mostrando até 20 valores únicos por coluna)")
    
    for i, col in enumerate(df_completo.columns, 1):
        print(f"\n   📊 Coluna {i}: '{col}'")
        print(f"      Tipo: {df_completo[col].dtype}")
        
        # Valores não nulos
        nao_nulos = df_completo[col].dropna()
        print(f"      Valores não nulos: {len(nao_nulos)} de {len(df_completo)}")
        
        if len(nao_nulos) > 0:
            # Valores únicos
            valores_unicos = nao_nulos.unique()
            print(f"      Valores únicos: {len(valores_unicos)}")
            
            # Mostrar amostra de valores únicos
            amostra = valores_unicos[:20].tolist()
            print(f"      Amostra (até 20):")
            for val in amostra:
                print(f"         • '{val}'")
            
            # Testar identificação de unidades nesta coluna
            unidades_encontradas = {}
            for val in valores_unicos[:100]:  # Testar até 100 valores únicos
                unidade = identificar_unidade(str(val))
                if unidade:
                    if unidade not in unidades_encontradas:
                        unidades_encontradas[unidade] = 0
                    unidades_encontradas[unidade] += 1
            
            if unidades_encontradas:
                print(f"      ✅ UNIDADES IDENTIFICADAS NESTA COLUNA:")
                for unidade, count in unidades_encontradas.items():
                    print(f"         • {unidade}: {count} ocorrências")
            else:
                print(f"      ❌ Nenhuma unidade identificada nesta coluna")
    
    return df_completo

def processar_arquivo_final(caminho, nome_arquivo, tipo):
    """Processa arquivo com estratégia adaptativa"""
    
    # Primeiro, diagnóstico completo
    df = super_diagnostico(caminho, nome_arquivo)
    if df is None:
        return {u: 0 for u in UNIDADES_ALVO}
    
    resultado = {u: 0 for u in UNIDADES_ALVO}
    
    print(f"\n🎯 ESTRATÉGIA DE PROCESSAMENTO PARA {tipo.upper()}:")
    
    if tipo == 'usuarios':
        print("   Buscando em colunas que contenham 'etiquetas' ou 'tags'")
        # Procurar colunas relacionadas a etiquetas/tags
        colunas_alvo = []
        for col in df.columns:
            col_lower = col.lower()
            if any(palavra in col_lower for palavra in ['etiqueta', 'tag', 'grupo', 'setor', 'unidade', 'empresa']):
                colunas_alvo.append(col)
                print(f"   ✅ Coluna candidata: '{col}'")
        
        if not colunas_alvo:
            print("   ⚠️ Nenhuma coluna específica encontrada, buscando em TODAS as colunas")
            colunas_alvo = df.columns.tolist()
        
        # Buscar unidades nas colunas alvo
        for col in colunas_alvo:
            valores = df[col].dropna().astype(str)
            for val in valores:
                unidade = identificar_unidade(val)
                if unidade and unidade in UNIDADES_ALVO:
                    resultado[unidade] += 1
    
    elif tipo == 'presencial':
        print("   Buscando em colunas que contenham 'empresa', 'unidade' ou 'cliente'")
        colunas_alvo = []
        for col in df.columns:
            col_lower = col.lower()
            if any(palavra in col_lower for palavra in ['empresa', 'unidade', 'cliente', 'razao', 'nome', 'fantasia']):
                colunas_alvo.append(col)
                print(f"   ✅ Coluna candidata: '{col}'")
        
        if not colunas_alvo:
            print("   ⚠️ Nenhuma coluna específica encontrada, buscando em TODAS as colunas")
            colunas_alvo = df.columns.tolist()
        
        # Buscar unidades nas colunas alvo
        for col in colunas_alvo:
            valores = df[col].dropna().astype(str)
            for val in valores:
                unidade = identificar_unidade(val)
                if unidade and unidade in UNIDADES_ALVO:
                    resultado[unidade] += 1
    
    elif tipo == 'conversas':
        print("   Buscando em colunas que contenham 'unidade'")
        colunas_alvo = []
        for col in df.columns:
            col_lower = col.lower()
            if 'unidade' in col_lower:
                colunas_alvo.append(col)
                print(f"   ✅ Coluna candidata: '{col}'")
        
        if not colunas_alvo:
            print("   ⚠️ Nenhuma coluna de unidade encontrada, buscando em TODAS as colunas")
            colunas_alvo = df.columns.tolist()
        
        for col in colunas_alvo:
            valores = df[col].dropna().astype(str)
            for val in valores:
                unidade = identificar_unidade(val)
                if unidade and unidade in UNIDADES_ALVO:
                    resultado[unidade] += 1
    
    # Mostrar resultado
    total = sum(resultado.values())
    print(f"\n📈 RESULTADO {tipo.upper()}: Total = {total}")
    if total > 0:
        for u, c in resultado.items():
            if c > 0:
                print(f"   {u}: {c}")
    else:
        print("   ⚠️ NENHUMA UNIDADE ENCONTRADA!")
        print("   💡 Verifique o diagnóstico acima para ver o conteúdo das colunas")
    
    return resultado

def processar_campanhas_final(caminho):
    """Processa campanhas com diagnóstico completo"""
    mkt_result = {u: 0 for u in UNIDADES_ALVO}
    util_result = {u: 0 for u in UNIDADES_ALVO}
    
    df = super_diagnostico(caminho, 'campanhas.csv')
    if df is None:
        return mkt_result, util_result
    
    print(f"\n🎯 PROCESSANDO CAMPANHAS:")
    
    # Encontrar coluna de unidade/departamento
    col_unidade = None
    for col in df.columns:
        col_lower = col.lower()
        if any(p in col_lower for p in ['departamento', 'unidade', 'empresa', 'setor']):
            col_unidade = col
            print(f"   ✅ Coluna unidade: '{col}'")
            break
    
    # Encontrar coluna de categoria
    col_categoria = None
    for col in df.columns:
        col_lower = col.lower()
        if 'categoria' in col_lower or ('whatsapp' in col_lower and 'tipo' in col_lower):
            col_categoria = col
            print(f"   ✅ Coluna categoria: '{col}'")
            break
    
    if not col_unidade:
        print("   ❌ Coluna de unidade não encontrada")
        # Tentar usar a primeira coluna que tenha valores de texto
        for col in df.columns:
            if df[col].dtype == 'object':
                col_unidade = col
                print(f"   🔄 Usando coluna alternativa: '{col}'")
                break
    
    if not col_categoria:
        print("   ❌ Coluna de categoria não encontrada")
        # Procurar qualquer coluna que contenha 'marketing' ou 'utility'
        for col in df.columns:
            valores = df[col].dropna().astype(str).str.lower()
            if valores.str.contains('marketing|utility', na=False).any():
                col_categoria = col
                print(f"   🔄 Usando coluna alternativa para categoria: '{col}'")
                break
    
    if col_unidade and col_categoria:
        for idx, row in df.iterrows():
            unidade = identificar_unidade(str(row[col_unidade])) if pd.notna(row[col_unidade]) else None
            categoria = normalizar(str(row[col_categoria])) if pd.notna(row[col_categoria]) else ''
            
            if unidade and unidade in UNIDADES_ALVO:
                if 'MARKETING' in categoria:
                    mkt_result[unidade] += 1
                elif 'UTILITY' in categoria:
                    util_result[unidade] += 1
    
    print(f"\n📈 Marketing: {sum(mkt_result.values())}")
    if sum(mkt_result.values()) > 0:
        for u, c in mkt_result.items():
            if c > 0:
                print(f"   {u}: {c}")
    
    print(f"\n📈 Utility: {sum(util_result.values())}")
    if sum(util_result.values()) > 0:
        for u, c in util_result.items():
            if c > 0:
                print(f"   {u}: {c}")
    
    return mkt_result, util_result

def formatar_excel_com_cores(writer, sheet_name, dados_percentuais, dados_valores):
    """Formata o Excel com cores e estilos"""
    AZUL_CABECALHO = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    AZUL_CLARO_LINHA = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    BRANCO = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    FONTE_BRANCA = Font(color="FFFFFF", bold=True, size=11)
    FONTE_PRETO = Font(color="000000", size=10)
    BORDA = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # ABA 1: PERCENTUAIS
    colunas_percentuais = [
        'Unidade',
        'Franquia (%)',
        'Mensagens Receptivas Excedentes (%)',
        'Licenças Excedentes (%)',
        'Atendimento Presencial (%)',
        'WhatsApp Marketing (%)',
        'WhatsApp Utility (%)',
        'Número Oficial WhatsApp (WABA) (%)'
    ]
    
    dados_numericos = []
    for item in dados_percentuais:
        dados_numericos.append([
            item['Unidade'],
            item['Franquia'],
            item['Mensagens Receptivas Excedentes'],
            item['Licenças Excedentes'],
            item['Atendimento Presencial'],
            item['WhatsApp Marketing'],
            item['WhatsApp Utility'],
            item['Número Oficial WhatsApp (WABA)']
        ])
    
    df_pct = pd.DataFrame(dados_numericos, columns=colunas_percentuais)
    df_pct.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
    
    ws = writer.sheets[sheet_name]
    num_data_rows = len(dados_numericos)
    total_row = num_data_rows + 2
    
    ws.cell(row=total_row, column=1).value = "TOTAIS:"
    for col in range(2, len(colunas_percentuais) + 1):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}2:{col_letter}{num_data_rows + 1})"
        ws.cell(row=total_row, column=col).value = formula
        ws.cell(row=total_row, column=col).number_format = '0.00"%"'
    
    for row in range(2, num_data_rows + 2):
        for col in range(2, len(colunas_percentuais) + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = '0.00"%"'
    
    for col in range(1, len(colunas_percentuais) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = AZUL_CABECALHO
        cell.font = FONTE_BRANCA
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDA
    
    for row in range(2, total_row + 1):
        for col in range(1, len(colunas_percentuais) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDA
            cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
            
            if row == total_row:
                cell.font = Font(bold=True, size=10, color="000000")
                cell.fill = BRANCO
            else:
                cell.font = FONTE_PRETO
                if (row - 1) % 2 == 0:
                    cell.fill = AZUL_CLARO_LINHA
    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)
    
    # ABA 2: VALORES EM R$
    colunas_valores = [
        'Unidade',
        'Franquia',
        'Mensagens Receptivas Excedentes',
        'Licenças Excedentes',
        'Atendimento Presencial',
        'WhatsApp Marketing',
        'WhatsApp Utility',
        'Número Oficial WhatsApp (WABA)',
        'Total por Unidade'
    ]
    
    dados_valores_numericos = []
    for item in dados_valores:
        dados_valores_numericos.append([
            item['Unidade'],
            item['Franquia'],
            item['Mensagens Receptivas Excedentes'],
            item['Licenças Excedentes'],
            item['Atendimento Presencial'],
            item['WhatsApp Marketing'],
            item['WhatsApp Utility'],
            item['Número Oficial WhatsApp (WABA)'],
            item['Total por Unidade']
        ])
    
    df_val = pd.DataFrame(dados_valores_numericos, columns=colunas_valores)
    df_val.to_excel(writer, sheet_name="Rateio de Valores (R$)", index=False, startrow=0)
    
    ws2 = writer.sheets["Rateio de Valores (R$)"]
    num_data_rows_val = len(dados_valores_numericos)
    total_row_val = num_data_rows_val + 2
    
    ws2.cell(row=total_row_val, column=1).value = "TOTAL:"
    for col in range(2, len(colunas_valores) + 1):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}2:{col_letter}{num_data_rows_val + 1})"
        ws2.cell(row=total_row_val, column=col).value = formula
        ws2.cell(row=total_row_val, column=col).number_format = '#,##0.00'
    
    for row in range(2, num_data_rows_val + 2):
        for col in range(2, len(colunas_valores) + 1):
            cell = ws2.cell(row=row, column=col)
            cell.number_format = '#,##0.00'
    
    for col in range(1, len(colunas_valores) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = AZUL_CABECALHO
        cell.font = FONTE_BRANCA
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDA
    
    for row in range(2, total_row_val + 1):
        for col in range(1, len(colunas_valores) + 1):
            cell = ws2.cell(row=row, column=col)
            cell.border = BORDA
            
            if col == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            if row == total_row_val:
                cell.font = Font(bold=True, size=10, color="000000")
                cell.fill = BRANCO
            else:
                cell.font = FONTE_PRETO
                if (row - 1) % 2 == 0:
                    cell.fill = AZUL_CLARO_LINHA
    
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws2.column_dimensions[col_letter].width = min(max_len + 3, 35)

# HTML da interface
HTML = '''
<!DOCTYPE html>
<html>
<head>
    <title>Rateio Norte Saneamento</title>
    <meta charset="UTF-8">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: 'Segoe UI', Arial, sans-serif; 
            background: linear-gradient(135deg, #1e3c72, #2a5298); 
            min-height: 100vh; 
            padding: 20px; 
        }
        .container { max-width: 1000px; margin: 0 auto; }
        .card { 
            background: white; 
            border-radius: 20px; 
            padding: 30px; 
            box-shadow: 0 10px 30px rgba(0,0,0,0.2); 
        }
        h1 { 
            color: #1e3c72; 
            text-align: center; 
            margin-bottom: 30px; 
            font-size: 24px;
        }
        .upload-grid { 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); 
            gap: 15px; 
            margin-bottom: 30px; 
        }
        .upload-item { 
            border: 2px dashed #ccc; 
            border-radius: 15px; 
            padding: 20px; 
            text-align: center; 
            cursor: pointer; 
            transition: all 0.3s; 
        }
        .upload-item:hover { 
            border-color: #2a5298; 
            background: #f0f4ff; 
            transform: translateY(-2px); 
        }
        .upload-item .icon { font-size: 40px; margin-bottom: 10px; }
        .upload-item h3 { margin-bottom: 5px; color: #333; }
        .upload-item p { font-size: 12px; color: #666; }
        .file-name { 
            font-size: 11px; 
            color: #28a745; 
            margin-top: 8px; 
            word-break: break-all; 
        }
        input { display: none; }
        .btn { 
            background: linear-gradient(135deg, #1e3c72, #2a5298); 
            color: white; 
            border: none; 
            padding: 12px 30px; 
            border-radius: 10px; 
            font-size: 16px; 
            cursor: pointer; 
            width: 100%; 
            font-weight: bold; 
            transition: all 0.3s; 
        }
        .btn:hover { 
            transform: translateY(-2px); 
            box-shadow: 0 5px 15px rgba(0,0,0,0.3); 
        }
        .loading { 
            display: none; 
            text-align: center; 
            padding: 20px; 
            margin-top: 20px; 
        }
        .spinner { 
            border: 4px solid #f3f3f3; 
            border-top: 4px solid #2a5298; 
            border-radius: 50%; 
            width: 40px; 
            height: 40px; 
            animation: spin 1s linear infinite; 
            margin: 0 auto; 
        }
        @keyframes spin { 
            0% { transform: rotate(0deg); } 
            100% { transform: rotate(360deg); } 
        }
        .status { 
            margin-top: 20px; 
            padding: 15px; 
            border-radius: 10px; 
        }
        .success { 
            background: #d4edda; 
            color: #155724; 
            border: 1px solid #c3e6cb; 
        }
        .error { 
            background: #f8d7da; 
            color: #721c24; 
            border: 1px solid #f5c6cb; 
        }
    </style>
</head>
<body>
<div class="container">
    <div class="card">
        <h1>📊 RATEIO NORTE SANEAMENTO - ABRIL/2026</h1>
        <div class="upload-grid">
            <div class="upload-item" onclick="document.getElementById('conversas').click()">
                <div class="icon">💬</div>
                <h3>Conversas</h3>
                <p>conversas.csv</p>
                <div class="file-name" id="conv-name"></div>
                <input type="file" id="conversas" accept=".csv">
            </div>
            <div class="upload-item" onclick="document.getElementById('campanhas').click()">
                <div class="icon">📢</div>
                <h3>Campanhas</h3>
                <p>campanhas.csv</p>
                <div class="file-name" id="camp-name"></div>
                <input type="file" id="campanhas" accept=".csv">
            </div>
            <div class="upload-item" onclick="document.getElementById('presencial').click()">
                <div class="icon">🏢</div>
                <h3>Presencial</h3>
                <p>presencial.csv</p>
                <div class="file-name" id="pres-name"></div>
                <input type="file" id="presencial" accept=".csv">
            </div>
            <div class="upload-item" onclick="document.getElementById('usuarios').click()">
                <div class="icon">👥</div>
                <h3>Usuários</h3>
                <p>usuarios.csv</p>
                <div class="file-name" id="user-name"></div>
                <input type="file" id="usuarios" accept=".csv">
            </div>
        </div>
        <button class="btn" onclick="processar()">🚀 GERAR RATEIO</button>
        <div id="loading" class="loading">
            <div class="spinner"></div>
            <p>Processando arquivos... Veja o diagnóstico no console!</p>
        </div>
        <div id="status"></div>
    </div>
</div>
<script>
    ['conversas', 'campanhas', 'presencial', 'usuarios'].forEach(function(id) {
        var input = document.getElementById(id);
        input.onchange = function(e) {
            var spanId;
            if (id === 'conversas') spanId = 'conv-name';
            else if (id === 'campanhas') spanId = 'camp-name';
            else if (id === 'presencial') spanId = 'pres-name';
            else if (id === 'usuarios') spanId = 'user-name';
            
            var span = document.getElementById(spanId);
            if (span && this.files[0]) {
                span.innerHTML = '✅ ' + this.files[0].name;
            }
        };
    });
    
    async function processar() {
        var formData = new FormData();
        var hasFiles = false;
        var ids = ['conversas', 'campanhas', 'presencial', 'usuarios'];
        
        for (var i = 0; i < ids.length; i++) {
            var file = document.getElementById(ids[i]).files[0];
            if (file) {
                formData.append(ids[i], file);
                hasFiles = true;
            }
        }
        
        if (!hasFiles) {
            alert('❌ Selecione pelo menos um arquivo CSV');
            return;
        }
        
        document.getElementById('loading').style.display = 'block';
        document.getElementById('status').innerHTML = '';
        
        try {
            var response = await fetch('/api/rateio', { 
                method: 'POST', 
                body: formData 
            });
            
            if (!response.ok) {
                throw new Error('Erro na requisição: ' + response.status);
            }
            
            var data = await response.json();
            
            if (data.success) {
                var link = document.createElement('a');
                link.href = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,' + data.excel;
                link.download = 'rateio_norte_' + new Date().toISOString().slice(0,19).replace(/:/g, '-') + '.xlsx';
                link.click();
                
                document.getElementById('status').innerHTML = '<div class="status success">' + data.mensagem + '</div>';
            } else {
                document.getElementById('status').innerHTML = '<div class="status error">❌ ' + data.erro + '</div>';
            }
        } catch(e) {
            document.getElementById('status').innerHTML = '<div class="status error">❌ Erro: ' + e.message + '</div>';
        }
        
        document.getElementById('loading').style.display = 'none';
    }
</script>
</body>
</html>
'''

@app.route('/')
def index():
    return HTML

@app.route('/api/rateio', methods=['POST'])
def rateio():
    caminhos = {}
    
    # Salvar arquivos enviados
    for nome in ['conversas', 'campanhas', 'presencial', 'usuarios']:
        if nome in request.files and request.files[nome].filename:
            caminho = os.path.join(UPLOAD_FOLDER, f"{nome}_{datetime.now().timestamp()}.csv")
            request.files[nome].save(caminho)
            caminhos[nome] = caminho
            print(f"\n📁 Arquivo {nome} salvo: {request.files[nome].filename} ({os.path.getsize(caminho)} bytes)")
    
    try:
        print("\n" + "="*80)
        print("🔬 SUPER DIAGNÓSTICO - ANALISANDO TODOS OS ARQUIVOS")
        print("="*80)
        
        # Processar cada arquivo com diagnóstico ultra-detalhado
        conversas = processar_arquivo_final(caminhos.get('conversas'), 'conversas.csv', 'conversas')
        usuarios = processar_arquivo_final(caminhos.get('usuarios'), 'usuarios.csv', 'usuarios')
        presencial = processar_arquivo_final(caminhos.get('presencial'), 'presencial.csv', 'presencial')
        
        # Campanhas
        if caminhos.get('campanhas'):
            marketing, utility = processar_campanhas_final(caminhos.get('campanhas'))
        else:
            marketing = {u: 0 for u in UNIDADES_ALVO}
            utility = {u: 0 for u in UNIDADES_ALVO}
        
        # Calcular totais
        total_conv = sum(conversas.values())
        total_user = sum(usuarios.values())
        total_pres = sum(presencial.values())
        total_mkt = sum(marketing.values())
        total_util = sum(utility.values())
        
        print(f"\n{'='*80}")
        print("📈 TOTAIS FINAIS:")
        print(f"  Conversas: {total_conv}")
        print(f"  Usuários (Licenças): {total_user}")
        print(f"  Presencial: {total_pres}")
        print(f"  WhatsApp Marketing: {total_mkt}")
        print(f"  WhatsApp Utility: {total_util}")
        
        # Calcular valores
        valor_total_mensagens = total_conv * VALORES_UNITARIOS['mensagens_receptivas']
        valor_total_licencas = total_user * VALORES_UNITARIOS['licencas']
        valor_total_presencial = total_pres * VALORES_UNITARIOS['atendimento_presencial']
        valor_total_marketing = total_mkt * VALORES_UNITARIOS['whatsapp_marketing']
        valor_total_utility = total_util * VALORES_UNITARIOS['whatsapp_utility']
        valor_total_waba = len(UNIDADES_ALVO) * VALORES_UNITARIOS['numero_oficial_waba']
        
        dados_percentuais = []
        dados_valores = []
        
        # Calcular rateio por unidade
        for unidade in UNIDADES_ALVO:
            pct_conv = round(conversas[unidade] / total_conv * 100, 2) if total_conv > 0 else 0
            pct_user = round(usuarios[unidade] / total_user * 100, 2) if total_user > 0 else 0
            pct_pres = round(presencial[unidade] / total_pres * 100, 2) if total_pres > 0 else 0
            pct_mkt = round(marketing[unidade] / total_mkt * 100, 2) if total_mkt > 0 else 0
            pct_util = round(utility[unidade] / total_util * 100, 2) if total_util > 0 else 0
            
            dados_percentuais.append({
                'Unidade': NOMES_PLANILHA[unidade],
                'Franquia': 10.00,
                'Mensagens Receptivas Excedentes': pct_conv,
                'Licenças Excedentes': pct_user,
                'Atendimento Presencial': pct_pres,
                'WhatsApp Marketing': pct_mkt,
                'WhatsApp Utility': pct_util,
                'Número Oficial WhatsApp (WABA)': 10.00
            })
            
            valor_mensagens = (pct_conv / 100) * valor_total_mensagens
            valor_licencas = (pct_user / 100) * valor_total_licencas
            valor_presencial_calc = (pct_pres / 100) * valor_total_presencial
            valor_marketing_calc = (pct_mkt / 100) * valor_total_marketing
            valor_utility_calc = (pct_util / 100) * valor_total_utility
            valor_waba = (10 / 100) * valor_total_waba
            
            total_unidade = (VALORES_UNITARIOS['franquia'] + valor_mensagens + valor_licencas + 
                           valor_presencial_calc + valor_marketing_calc + valor_utility_calc + valor_waba)
            
            dados_valores.append({
                'Unidade': NOMES_PLANILHA[unidade],
                'Franquia': VALORES_UNITARIOS['franquia'],
                'Mensagens Receptivas Excedentes': round(valor_mensagens, 2),
                'Licenças Excedentes': round(valor_licencas, 2),
                'Atendimento Presencial': round(valor_presencial_calc, 2),
                'WhatsApp Marketing': round(valor_marketing_calc, 2),
                'WhatsApp Utility': round(valor_utility_calc, 2),
                'Número Oficial WhatsApp (WABA)': round(valor_waba, 2),
                'Total por Unidade': round(total_unidade, 2)
            })
        
        # Gerar Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            formatar_excel_com_cores(writer, "Percentual de Consumo por Item", dados_percentuais, dados_valores)
        
        output.seek(0)
        
        mensagem = f'✅ Rateio gerado! Conversas: {total_conv} | Usuários: {total_user} | Presencial: {total_pres} | Marketing: {total_mkt} | Utility: {total_util}'
        print(f"\n{mensagem}")
        print(f"{'='*80}\n")
        
        return jsonify({
            'success': True,
            'excel': base64.b64encode(output.getvalue()).decode('utf-8'),
            'mensagem': mensagem
        })
        
    except Exception as e:
        print(f"\n❌ ERRO: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': f'Erro: {str(e)}'}), 500
        
    finally:
        for p in caminhos.values():
            if os.path.exists(p):
                try:
                    os.remove(p)
                except:
                    pass
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)